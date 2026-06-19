import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from cexie_excel_tool.processor import (
    CorrectionRule,
    apply_corrections_to_folder,
    copy_rules_to_all_points,
    discover_workbooks,
    export_a0_summary,
    load_rules,
    save_rules,
)


def make_workbook(path: Path, sheet_name: str = "手动测斜-18T-ZQT05-202606060806") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["深度（m）", "A0", "A180", "设备编号"])
    ws.append([0.5, -22.5, 23.0, "CM024"])
    ws.append([1.0, None, "bad", None])
    ws.column_dimensions["A"].width = 18
    wb.save(path)


def make_full_inclinometer_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "手动测斜-18T-ZQT03-202606070749"
    ws.append(["深度（m）", "A0", "A180", "A轴管口起算", "A轴管底起算", "B0", "B180", "B轴管口起算", "B轴管底起算", "设备编号"])
    ws.append([0.5, -2, 4, 999, 999, 10, 4, 999, 999, "CM024"])
    ws.append([1.0, -1, 5, 999, 999, 9, 3, 999, 999, None])
    ws.append([1.5, 0, 6, 999, 999, 8, 2, 999, 999, None])
    wb.save(path)


def make_multi_sheet_inclinometer_workbook(path: Path) -> None:
    wb = Workbook()
    for index, title in enumerate(["手动测斜-18T-ZQT03-202606110750", "手动测斜2-JM3-202606110652"]):
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = title
        ws.append(["深度（m）", "A0", "A180", "A轴管口起算", "A轴管底起算", "B0", "B180", "B轴管口起算", "B轴管底起算", "设备编号"])
        ws.append([0.5, -23.03, 23.72, None, None, 5.51, -26.67, None, None, "CM024290037"])
        ws.append([1.0, -27.16, 29.75, None, None, -5.62, -15.06, None, None, None])
    wb.save(path)


def make_unsorted_multi_sheet_workbook(path: Path) -> None:
    wb = Workbook()
    sheets = [
        ("手动测斜-18T-ZQT10-202606110750", -10),
        ("手动测斜2-NP-ZQT6-202606110627", -6),
        ("手动测斜-18T-ZQT03-202606110750", -3),
        ("手动测斜2-NP-ZQT3-202606110643", -30),
    ]
    for index, (title, a0_value) in enumerate(sheets):
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = title
        ws.append(["深度（m）", "A0", "A180"])
        ws.append([0.5, a0_value, 0])
    wb.save(path)


class ExcelProcessorTests(unittest.TestCase):
    def test_discovers_xlsx_files_headers_and_point_ids(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            make_workbook(folder / "手动测斜-18T-ZQT05-原始数据.xlsx")
            (folder / "notes.txt").write_text("ignore", encoding="utf-8")

            found = discover_workbooks(folder)

            self.assertEqual(len(found), 1)
            self.assertEqual(found[0].point_id, "18T-ZQT05")
            self.assertEqual(found[0].headers, ["深度（m）", "A0", "A180", "设备编号"])

    def test_discovers_each_sheet_in_multi_sheet_workbooks(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            make_multi_sheet_inclinometer_workbook(folder / "手动-2026-06-11-单日所有原始数据.xlsx")

            found = discover_workbooks(folder)

            self.assertEqual([item.point_id for item in found], ["18T-ZQT03", "JM3"])
            self.assertEqual([item.sheet_name for item in found], ["手动测斜-18T-ZQT03-202606110750", "手动测斜2-JM3-202606110652"])

    def test_applies_numeric_and_replace_rules_without_overwriting_source(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            source = folder / "手动测斜-18T-ZQT05-原始数据.xlsx"
            make_workbook(source)
            rules = {
                "18T-ZQT05": {
                    "深度（m）": CorrectionRule("add", "1"),
                    "A0": CorrectionRule("mul", "2"),
                    "A180": CorrectionRule("div", "2"),
                    "设备编号": CorrectionRule("replace", "DEVICE-X"),
                }
            }

            summary = apply_corrections_to_folder(folder, rules)

            self.assertEqual(summary.processed_files, 1)
            self.assertEqual(summary.skipped_cells, 1)
            self.assertTrue(source.exists())
            self.assertEqual(len(summary.outputs), 1)
            out = summary.outputs[0]
            self.assertIn("corrected", str(out))
            self.assertIn("修正后", out.name)

            wb = load_workbook(out, data_only=True)
            ws = wb.active
            self.assertEqual(ws["A2"].value, 1.5)
            self.assertEqual(ws["B2"].value, -45.0)
            self.assertEqual(ws["C2"].value, 11.5)
            self.assertEqual(ws["D2"].value, "DEVICE-X")
            self.assertIsNone(ws["B3"].value)
            self.assertEqual(ws["C3"].value, "bad")
            self.assertEqual(ws.column_dimensions["A"].width, 18)

            original = load_workbook(source, data_only=True).active
            self.assertEqual(original["A2"].value, 0.5)
            self.assertEqual(original["D2"].value, "CM024")

    def test_formats_generated_numeric_cells_to_one_or_two_decimals(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            source = folder / "手动测斜-18T-ZQT05-原始数据.xlsx"
            make_workbook(source)
            rules = {
                "18T-ZQT05": {
                    "深度（m）": CorrectionRule("formula", "x + 0.004"),
                    "A0": CorrectionRule("formula", "x / 7"),
                    "A180": CorrectionRule("formula", "x / 2"),
                }
            }

            summary = apply_corrections_to_folder(folder, rules)

            wb = load_workbook(summary.outputs[0], data_only=False)
            ws = wb.active
            self.assertEqual(ws["A2"].value, 0.5)
            self.assertEqual(ws["A2"].number_format, "General")
            self.assertEqual(ws["A3"].value, 1)
            self.assertEqual(ws["A3"].number_format, "General")
            self.assertEqual(ws["B2"].value, -3.21)
            self.assertEqual(ws["B2"].number_format, "General")
            self.assertEqual(ws["C2"].value, 11.5)
            self.assertEqual(ws["C2"].number_format, "General")
            self.assertEqual(ws["D2"].value, "CM024")
            wb.close()

    def test_generates_unique_output_name_when_file_exists(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            make_workbook(folder / "手动测斜-18T-ZQT05-原始数据.xlsx")
            corrected = folder / "corrected"
            corrected.mkdir()
            existing = corrected / "手动测斜-18T-ZQT05-原始数据-修正后.xlsx"
            existing.write_text("existing", encoding="utf-8")

            summary = apply_corrections_to_folder(folder, {"18T-ZQT05": {"A0": CorrectionRule("add", "1")}})

            self.assertEqual(summary.processed_files, 1)
            self.assertNotEqual(summary.outputs[0], existing)
            self.assertTrue(summary.outputs[0].name.startswith("手动测斜-18T-ZQT05-原始数据-修正后-"))

    def test_rejects_division_by_zero(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            make_workbook(folder / "手动测斜-18T-ZQT05-原始数据.xlsx")

            with self.assertRaises(ValueError):
                apply_corrections_to_folder(folder, {"18T-ZQT05": {"A0": CorrectionRule("div", "0")}})

    def test_applies_formula_rule_with_x_and_rand_function(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            source = folder / "手动测斜-18T-ZQT05-原始数据.xlsx"
            make_workbook(source)
            rules = {
                "18T-ZQT05": {
                    "深度（m）": CorrectionRule("formula", "x * 2 + 1"),
                    "A0": CorrectionRule("formula", "x + rand() * 10"),
                }
            }

            with patch("cexie_excel_tool.processor.random.random", return_value=0.25):
                summary = apply_corrections_to_folder(folder, rules)

            self.assertEqual(summary.processed_files, 1)
            self.assertEqual(summary.skipped_cells, 0)
            ws = load_workbook(summary.outputs[0], data_only=True).active
            self.assertEqual(ws["A2"].value, 2.0)
            self.assertEqual(ws["A3"].value, 3.0)
            self.assertEqual(ws["B2"].value, -20.0)
            self.assertIsNone(ws["B3"].value)

    def test_formula_rule_can_reference_same_row_source_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            source = folder / "手动测斜-18T-ZQT05-原始数据.xlsx"
            make_workbook(source)
            rules = {
                "18T-ZQT05": {
                    "A0": CorrectionRule("add", "100"),
                    "A180": CorrectionRule("formula", "x + [A0] + [深度（m）]"),
                }
            }

            summary = apply_corrections_to_folder(folder, rules)

            self.assertEqual(summary.processed_files, 1)
            self.assertEqual(summary.skipped_cells, 1)
            ws = load_workbook(summary.outputs[0], data_only=True).active
            self.assertEqual(ws["B2"].value, 77.5)
            self.assertEqual(ws["C2"].value, 1.0)
            self.assertEqual(ws["C3"].value, "bad")

    def test_formula_rule_rejects_missing_source_column_reference(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            make_workbook(folder / "手动测斜-18T-ZQT05-原始数据.xlsx")

            with self.assertRaises(ValueError):
                apply_corrections_to_folder(folder, {"18T-ZQT05": {"A0": CorrectionRule("formula", "x + [不存在列]")}})

    def test_recalculates_inclinometer_derived_columns_after_raw_adjustments(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            source = folder / "手动测斜-18T-ZQT03-原始数据.xlsx"
            make_full_inclinometer_workbook(source)
            rules = {
                "18T-ZQT03": {
                    "A0": CorrectionRule("add", "1"),
                    "B180": CorrectionRule("formula", "x + [深度（m）]"),
                }
            }

            summary = apply_corrections_to_folder(folder, rules)

            self.assertEqual(summary.processed_files, 1)
            ws = load_workbook(summary.outputs[0], data_only=True).active
            self.assertEqual([ws["D2"].value, ws["D3"].value, ws["D4"].value], [2.5, 5.0, 7.5])
            self.assertEqual([ws["E2"].value, ws["E3"].value, ws["E4"].value], [-7.5, -5.0, -2.5])
            self.assertEqual([ws["H2"].value, ws["H3"].value, ws["H4"].value], [-2.75, -5.25, -7.5])
            self.assertEqual([ws["I2"].value, ws["I3"].value, ws["I4"].value], [7.5, 4.75, 2.25])

    def test_inclinometer_recalculation_rounds_each_increment_half_up_before_accumulating(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            source = folder / "手动测斜-18T-ZQT03-原始数据.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "手动测斜-18T-ZQT03-202606070749"
            ws.append(["深度（m）", "A0", "A180", "A轴管口起算", "A轴管底起算", "B0", "B180", "B轴管口起算", "B轴管底起算", "设备编号"])
            ws.append([0.5, -23.03, 23.72, None, None, 5.51, -26.67, None, None, "CM024290037"])
            ws.append([1.0, -27.16, 29.75, None, None, -5.62, -15.06, None, None, None])
            wb.save(source)

            summary = apply_corrections_to_folder(folder, {"18T-ZQT03": {"A0": CorrectionRule("add", "0")}})

            ws = load_workbook(summary.outputs[0], data_only=True).active
            self.assertEqual(ws["D2"].value, 23.38)
            self.assertEqual(ws["D3"].value, 51.84)
            self.assertEqual(ws["E2"].value, -51.84)
            self.assertEqual(ws["E3"].value, -28.46)
            self.assertEqual(ws["H2"].value, -16.09)
            self.assertEqual(ws["H3"].value, -20.81)
            self.assertEqual(ws["I2"].value, 20.81)
            self.assertEqual(ws["I3"].value, 4.72)

    @patch("cexie_excel_tool.processor.random.randint", side_effect=[14, 35, 15, 6])
    def test_processes_all_sheets_in_a_multi_sheet_workbook(self, _mock_randint):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            source = folder / "手动-2026-06-11-单日所有原始数据.xlsx"
            make_multi_sheet_inclinometer_workbook(source)
            rules = {
                "18T-ZQT03": {"A0": CorrectionRule("add", "0")},
                "JM3": {"A0": CorrectionRule("add", "1")},
            }

            summary = apply_corrections_to_folder(folder, rules)

            self.assertEqual(summary.processed_files, 1)
            self.assertEqual(len(summary.outputs), 1)
            wb = load_workbook(summary.outputs[0], data_only=True)
            first = wb["手动测斜-18T-ZQT03-202606111435"]
            second = wb["手动测斜2-JM3-202606111506"]
            self.assertEqual(first["D3"].value, 51.84)
            self.assertEqual(first["E2"].value, -51.84)
            self.assertEqual(second["B2"].value, -22.03)
            self.assertEqual(second["D2"].value, 22.88)
            self.assertEqual(second["D3"].value, 50.84)
            self.assertEqual(second["E2"].value, -50.84)
            wb.close()

    @patch("cexie_excel_tool.processor.random.randint", side_effect=[14, 35, 15, 6])
    def test_generated_workbook_renames_sheet_time_to_random_afternoon_time(self, _mock_randint):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            source = folder / "手动-2026-06-11-单日所有原始数据.xlsx"
            make_multi_sheet_inclinometer_workbook(source)

            summary = apply_corrections_to_folder(
                folder,
                {
                    "18T-ZQT03": {"A0": CorrectionRule("add", "0")},
                    "JM3": {"A0": CorrectionRule("add", "0")},
                },
            )

            wb = load_workbook(summary.outputs[0], read_only=True)
            sheetnames = wb.sheetnames
            wb.close()
            self.assertEqual(sheetnames, ["手动测斜-18T-ZQT03-202606111435", "手动测斜2-JM3-202606111506"])

    def test_exports_a0_summary_from_all_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            source = folder / "手动-2026-06-11-单日所有原始数据.xlsx"
            make_multi_sheet_inclinometer_workbook(source)

            output = export_a0_summary(source)

            self.assertEqual(output.name, "手动-2026-06-11-单日所有原始数据-A0汇总.xlsx")
            wb = load_workbook(output, data_only=True)
            ws = wb.active
            self.assertEqual(ws.title, "A0汇总")
            self.assertEqual([ws["A1"].value, ws["B1"].value, ws["C1"].value], ["深度（m）", "18T-ZQT03", "JM3"])
            self.assertEqual([ws["A2"].value, ws["B2"].value, ws["C2"].value], [0.5, -23.03, -23.03])
            self.assertEqual([ws["A3"].value, ws["B3"].value, ws["C3"].value], [1.0, -27.16, -27.16])
            self.assertEqual(ws.freeze_panes, "B2")
            wb.close()

    def test_exports_a0_summary_sorted_by_point_id_not_sheet_order(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            source = folder / "手动-2026-06-11-单日所有原始数据.xlsx"
            make_unsorted_multi_sheet_workbook(source)

            output = export_a0_summary(source)

            wb = load_workbook(output, data_only=True)
            ws = wb.active
            self.assertEqual(
                [ws.cell(row=1, column=column).value for column in range(1, ws.max_column + 1)],
                ["深度（m）", "18T-ZQT03", "18T-ZQT10", "NP-ZQT3", "NP-ZQT6"],
            )
            self.assertEqual(
                [ws.cell(row=2, column=column).value for column in range(1, ws.max_column + 1)],
                [0.5, -3, -10, -30, -6],
            )
            wb.close()

    def test_exports_a0_summary_with_unique_name_when_output_exists(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            source = folder / "手动-2026-06-11-单日所有原始数据.xlsx"
            make_multi_sheet_inclinometer_workbook(source)
            existing = folder / "手动-2026-06-11-单日所有原始数据-A0汇总.xlsx"
            existing.write_text("existing", encoding="utf-8")

            output = export_a0_summary(source)

            self.assertNotEqual(output, existing)
            self.assertTrue(output.name.startswith("手动-2026-06-11-单日所有原始数据-A0汇总-"))
            self.assertTrue(output.exists())

    def test_rejects_unsafe_formula_rule(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            make_workbook(folder / "手动测斜-18T-ZQT05-原始数据.xlsx")

            with self.assertRaises(ValueError):
                apply_corrections_to_folder(folder, {"18T-ZQT05": {"A0": CorrectionRule("formula", "__import__('os')")}})

    def test_saves_and_loads_rules(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rules.json"
            rules = {"18T-ZQT05": {"A0": CorrectionRule("add", "1.2"), "设备编号": CorrectionRule("replace", "X")}}

            save_rules(path, rules)
            raw = json.loads(path.read_text(encoding="utf-8"))
            loaded = load_rules(path)

            self.assertEqual(raw["version"], 1)
            self.assertEqual(loaded["18T-ZQT05"]["A0"], CorrectionRule("add", "1.2"))
            self.assertEqual(loaded["18T-ZQT05"]["设备编号"], CorrectionRule("replace", "X"))

    def test_copies_one_points_rules_to_all_matching_columns(self):
        workbooks = [
            make_info("18T-ZQT05", ["A0", "A180", "设备编号"]),
            make_info("18T-ZQT06", ["A0", "A180", "设备编号"]),
            make_info("18T-ZQT07", ["A0", "设备编号"]),
        ]
        rules = {
            "18T-ZQT05": {
                "A0": CorrectionRule("add", "1"),
                "A180": CorrectionRule("sub", "2"),
                "设备编号": CorrectionRule("replace", "DEVICE-X"),
            },
            "18T-ZQT06": {"A0": CorrectionRule("mul", "9")},
        }

        updated = copy_rules_to_all_points("18T-ZQT05", workbooks, rules)

        self.assertEqual(updated["18T-ZQT06"]["A0"], CorrectionRule("add", "1"))
        self.assertEqual(updated["18T-ZQT06"]["A180"], CorrectionRule("sub", "2"))
        self.assertEqual(updated["18T-ZQT06"]["设备编号"], CorrectionRule("replace", "DEVICE-X"))
        self.assertEqual(updated["18T-ZQT07"]["A0"], CorrectionRule("add", "1"))
        self.assertNotIn("A180", updated["18T-ZQT07"])
        self.assertEqual(updated["18T-ZQT07"]["设备编号"], CorrectionRule("replace", "DEVICE-X"))


def make_info(point_id: str, headers: list[str]):
    from cexie_excel_tool.processor import WorkbookInfo

    return WorkbookInfo(
        path=Path(f"{point_id}.xlsx"),
        point_id=point_id,
        sheet_name=point_id,
        headers=headers,
    )


if __name__ == "__main__":
    unittest.main()
