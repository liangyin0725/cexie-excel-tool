from __future__ import annotations

import ast
import json
import random
import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook


POINT_ID_RE = re.compile(r"\b\d+[A-Z]?-ZQTS?\d+\b", re.IGNORECASE)
COLUMN_REF_RE = re.compile(r"\[([^\]]+)\]")
VALID_OPS = {"add", "sub", "mul", "div", "replace", "formula"}
DERIVED_COLUMN_GROUPS = (
    ("A0", "A180", "A轴管口起算", "A轴管底起算"),
    ("B0", "B180", "B轴管口起算", "B轴管底起算"),
)


@dataclass(frozen=True)
class CorrectionRule:
    op: str
    value: str


@dataclass(frozen=True)
class WorkbookInfo:
    path: Path
    point_id: str
    sheet_name: str
    headers: list[str]


@dataclass(frozen=True)
class ProcessSummary:
    processed_files: int
    outputs: list[Path]
    skipped_cells: int
    skipped_files: list[str]


RuleMap = dict[str, dict[str, CorrectionRule]]


def discover_workbooks(folder: str | Path) -> list[WorkbookInfo]:
    root = Path(folder)
    if not root.exists() or not root.is_dir():
        raise ValueError(f"文件夹不存在: {root}")

    workbooks: list[WorkbookInfo] = []
    for path in sorted(root.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        headers = ["" if cell.value is None else str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        point_id = _extract_point_id(path.name) or _extract_point_id(ws.title) or path.stem
        workbooks.append(WorkbookInfo(path=path, point_id=point_id, sheet_name=ws.title, headers=headers))
        wb.close()
    return workbooks


def apply_corrections_to_folder(folder: str | Path, rules: Mapping[str, Mapping[str, CorrectionRule]]) -> ProcessSummary:
    root = Path(folder)
    output_dir = root / "corrected"
    output_dir.mkdir(exist_ok=True)

    outputs: list[Path] = []
    skipped_files: list[str] = []
    skipped_cells = 0

    for info in discover_workbooks(root):
        point_rules = dict(rules.get(info.point_id, {}))
        if not _has_effective_rules(point_rules):
            skipped_files.append(info.path.name)
            continue

        _validate_rules(point_rules)
        wb = load_workbook(info.path)
        ws = wb.worksheets[0]
        header_to_column = {
            "" if cell.value is None else str(cell.value): cell.column
            for cell in ws[1]
        }
        original_rows = {
            row: {
                header: ws.cell(row=row, column=column).value
                for header, column in header_to_column.items()
            }
            for row in range(2, ws.max_row + 1)
        }

        for header, rule in point_rules.items():
            if not rule.op or header not in header_to_column:
                continue
            if rule.op == "formula":
                _validate_formula_references(rule.value, header_to_column)
            column_index = header_to_column[header]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=column_index)
                new_value, skipped = _apply_rule(cell.value, rule, original_rows[row])
                if skipped:
                    skipped_cells += 1
                    continue
                cell.value = new_value

        skipped_cells += _recalculate_inclinometer_columns(ws, header_to_column)

        output_path = _unique_output_path(output_dir, info.path)
        wb.save(output_path)
        wb.close()
        outputs.append(output_path)

    return ProcessSummary(
        processed_files=len(outputs),
        outputs=outputs,
        skipped_cells=skipped_cells,
        skipped_files=skipped_files,
    )


def copy_rules_to_all_points(
    source_point_id: str,
    workbooks: list[WorkbookInfo],
    rules: Mapping[str, Mapping[str, CorrectionRule]],
) -> RuleMap:
    source_rules = dict(rules.get(source_point_id, {}))
    copied: RuleMap = {
        point_id: dict(column_rules)
        for point_id, column_rules in rules.items()
    }

    for info in workbooks:
        point_rules: dict[str, CorrectionRule] = {}
        available_headers = set(info.headers)
        for header, rule in source_rules.items():
            if header in available_headers:
                point_rules[header] = rule
        copied[info.point_id] = point_rules

    return copied


def save_rules(path: str | Path, rules: Mapping[str, Mapping[str, CorrectionRule]]) -> None:
    payload = {
        "version": 1,
        "points": {
            point_id: {
                header: {"op": rule.op, "value": rule.value}
                for header, rule in column_rules.items()
                if rule.op and rule.value != ""
            }
            for point_id, column_rules in rules.items()
        },
    }
    Path(path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_rules(path: str | Path) -> RuleMap:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if payload.get("version") != 1:
        raise ValueError("不支持的参数文件版本")
    loaded: RuleMap = {}
    for point_id, column_rules in payload.get("points", {}).items():
        loaded[point_id] = {}
        for header, raw_rule in column_rules.items():
            rule = CorrectionRule(str(raw_rule.get("op", "")), str(raw_rule.get("value", "")))
            if rule.op:
                _validate_rule(rule)
                loaded[point_id][header] = rule
    return loaded


def _extract_point_id(text: str) -> str | None:
    match = POINT_ID_RE.search(text)
    return match.group(0).upper() if match else None


def _has_effective_rules(rules: Mapping[str, CorrectionRule]) -> bool:
    return any(rule.op and rule.value != "" for rule in rules.values())


def _validate_rules(rules: Mapping[str, CorrectionRule]) -> None:
    for rule in rules.values():
        if rule.op:
            _validate_rule(rule)


def _validate_rule(rule: CorrectionRule) -> None:
    if rule.op not in VALID_OPS:
        raise ValueError(f"不支持的操作: {rule.op}")
    if rule.op == "formula":
        _validate_formula(rule.value)
    elif rule.op != "replace":
        number = _as_number(rule.value)
        if number is None:
            raise ValueError(f"数值规则必须填写数字: {rule.value}")
        if rule.op == "div" and number == 0:
            raise ValueError("除法规则不能填写 0")


def _apply_rule(value: object, rule: CorrectionRule, row_values: Mapping[str, object] | None = None) -> tuple[object, bool]:
    if value is None:
        return value, False
    if rule.op == "replace":
        return rule.value, False

    original = _as_number(value)
    if original is None:
        return value, True

    if rule.op == "formula":
        result = _evaluate_formula(rule.value, original, row_values or {})
        if result is None:
            return value, True
        return result, False

    operand = _as_number(rule.value)
    if operand is None:
        return value, True

    if rule.op == "add":
        return original + operand, False
    if rule.op == "sub":
        return original - operand, False
    if rule.op == "mul":
        return original * operand, False
    if rule.op == "div":
        return original / operand, False
    return value, True


def _recalculate_inclinometer_columns(ws, header_to_column: Mapping[str, int]) -> int:
    skipped = 0
    for zero_header, reverse_header, top_header, bottom_header in DERIVED_COLUMN_GROUPS:
        required = (zero_header, reverse_header, top_header, bottom_header)
        if any(header not in header_to_column for header in required):
            continue

        zero_col = header_to_column[zero_header]
        reverse_col = header_to_column[reverse_header]
        top_col = header_to_column[top_header]
        bottom_col = header_to_column[bottom_header]

        increments: list[Decimal | None] = []
        for row in range(2, ws.max_row + 1):
            increment = _half_difference_rounded(ws.cell(row=row, column=zero_col).value, ws.cell(row=row, column=reverse_col).value)
            if increment is None:
                increments.append(None)
                skipped += 2
            else:
                increments.append(increment)

        running_top = Decimal("0")
        for offset, increment in enumerate(increments):
            row = offset + 2
            if increment is None:
                continue
            running_top += increment
            ws.cell(row=row, column=top_col).value = _decimal_to_cell_value(running_top)

        running_bottom = Decimal("0")
        bottom_values: list[Decimal | None] = [None] * len(increments)
        for offset in range(len(increments) - 1, -1, -1):
            increment = increments[offset]
            if increment is None:
                continue
            running_bottom += increment
            bottom_values[offset] = -running_bottom

        for offset, value in enumerate(bottom_values):
            if value is not None:
                ws.cell(row=offset + 2, column=bottom_col).value = _decimal_to_cell_value(value)

    return skipped


def _half_difference_rounded(zero_value: object, reverse_value: object) -> Decimal | None:
    zero = _as_decimal(zero_value)
    reverse = _as_decimal(reverse_value)
    if zero is None or reverse is None:
        return None
    return ((reverse - zero) / Decimal("2")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _as_decimal(value: object) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, AttributeError):
        return None


def _decimal_to_cell_value(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _as_number(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _validate_formula(expression: str) -> None:
    prepared, _references = _prepare_formula(expression)
    try:
        tree = ast.parse(prepared, mode="eval")
    except SyntaxError as exc:
        raise ValueError(f"公式格式错误: {expression}") from exc
    _validate_formula_node(tree)


def _validate_formula_references(expression: str, header_to_column: Mapping[str, int]) -> None:
    _prepared, references = _prepare_formula(expression)
    missing = [header for header in references.values() if header not in header_to_column]
    if missing:
        raise ValueError(f"公式引用了不存在的列: {', '.join(missing)}")


def _evaluate_formula(expression: str, x: float, row_values: Mapping[str, object]) -> float | None:
    prepared, references = _prepare_formula(expression)
    ref_values: dict[str, float] = {}
    for variable, header in references.items():
        value = _as_number(row_values.get(header))
        if value is None:
            return None
        ref_values[variable] = value

    tree = ast.parse(prepared, mode="eval")
    _validate_formula_node(tree)
    return float(_eval_formula_node(tree.body, x, ref_values))


def _prepare_formula(expression: str) -> tuple[str, dict[str, str]]:
    references: dict[str, str] = {}

    def replace(match: re.Match[str]) -> str:
        header = match.group(1).strip()
        variable = f"col_{len(references)}"
        references[variable] = header
        return variable

    return COLUMN_REF_RE.sub(replace, expression), references


def _validate_formula_node(node: ast.AST) -> None:
    allowed_binary = (ast.Add, ast.Sub, ast.Mult, ast.Div)
    allowed_unary = (ast.UAdd, ast.USub)

    if isinstance(node, ast.Expression):
        _validate_formula_node(node.body)
    elif isinstance(node, ast.BinOp):
        if not isinstance(node.op, allowed_binary):
            raise ValueError("公式只支持 +、-、*、/ 四则运算")
        _validate_formula_node(node.left)
        _validate_formula_node(node.right)
    elif isinstance(node, ast.UnaryOp):
        if not isinstance(node.op, allowed_unary):
            raise ValueError("公式只支持正负号")
        _validate_formula_node(node.operand)
    elif isinstance(node, ast.Constant):
        if not isinstance(node.value, int | float):
            raise ValueError("公式里只能使用数字")
    elif isinstance(node, ast.Name):
        if node.id.lower() != "x" and not re.fullmatch(r"col_\d+", node.id):
            raise ValueError("公式里只能使用 x 或 [列名] 表示单元格数值")
    elif isinstance(node, ast.Call):
        if not isinstance(node.func, ast.Name) or node.func.id.lower() != "rand":
            raise ValueError("公式里只支持 rand() 随机数函数")
        if node.args or node.keywords:
            raise ValueError("rand() 不接受参数")
    else:
        raise ValueError("公式包含不支持的内容")


def _eval_formula_node(node: ast.AST, x: float, ref_values: Mapping[str, float]) -> float:
    if isinstance(node, ast.BinOp):
        left = _eval_formula_node(node.left, x, ref_values)
        right = _eval_formula_node(node.right, x, ref_values)
        if isinstance(node.op, ast.Add):
            return left + right
        if isinstance(node.op, ast.Sub):
            return left - right
        if isinstance(node.op, ast.Mult):
            return left * right
        if isinstance(node.op, ast.Div):
            return left / right
    if isinstance(node, ast.UnaryOp):
        value = _eval_formula_node(node.operand, x, ref_values)
        if isinstance(node.op, ast.UAdd):
            return value
        if isinstance(node.op, ast.USub):
            return -value
    if isinstance(node, ast.Constant) and isinstance(node.value, int | float):
        return float(node.value)
    if isinstance(node, ast.Name) and node.id.lower() == "x":
        return x
    if isinstance(node, ast.Name) and node.id in ref_values:
        return ref_values[node.id]
    if isinstance(node, ast.Call) and isinstance(node.func, ast.Name) and node.func.id.lower() == "rand":
        return random.random()
    raise ValueError("公式包含不支持的内容")


def _unique_output_path(output_dir: Path, source_path: Path) -> Path:
    target = output_dir / f"{source_path.stem}-修正后{source_path.suffix}"
    if not target.exists():
        return target
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    target = output_dir / f"{source_path.stem}-修正后-{stamp}{source_path.suffix}"
    counter = 2
    while target.exists():
        target = output_dir / f"{source_path.stem}-修正后-{stamp}-{counter}{source_path.suffix}"
        counter += 1
    return target
